# -----------------------------------
# REQUIREMENTS
# pip install streamlit yfinance pandas numpy openpyxl tqdm requests
#
# RUN:
#   streamlit run app.py
# -----------------------------------

import io
import requests
import numpy as np
import pandas as pd
import streamlit as st
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor
from openpyxl.utils import get_column_letter
import yfinance as yf

# ───────────────────────────────────────────────────────────────
# PAGE CONFIG
# ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Stock Screener",
    page_icon="📈",
    layout="wide",
)

st.title("📈 Hybrid Stock Screener")
st.caption("Screens S&P 500, NYSE, or NASDAQ for individual companies — ETFs, funds, and index products excluded.")

# ───────────────────────────────────────────────────────────────
# METRIC CONFIG  — name, score weight, description
# ───────────────────────────────────────────────────────────────
METRICS = {
    "OE_Yield":       {"label": "OE Yield",         "weight": 3, "desc": "Owner Earnings / Market Cap  (×3)"},
    "ROIC":           {"label": "ROIC",              "weight": 2, "desc": "Return on Invested Capital   (×2)"},
    "ROIC_Trend":     {"label": "ROIC Trend",        "weight": 2, "desc": "YoY change in ROIC           (×2)"},
    "RevenueGrowth":  {"label": "Revenue Growth",    "weight": 1, "desc": "Revenue growth rate          (×1)"},
    "EarningsGrowth": {"label": "Earnings Growth",   "weight": 1, "desc": "Earnings growth rate         (×1)"},
    "Piotroski":      {"label": "Piotroski Score",   "weight": 1, "desc": "Accounting health 0–9        (×1)"},
    "BuyingPressure": {"label": "Buying Pressure",   "weight": 2, "desc": "OBV + MFI + PCV composite    (×2)"},
}

ALL_SECTORS = [
    "All Sectors",
    "Communication Services",
    "Consumer Discretionary",
    "Consumer Staples",
    "Energy",
    "Financials",
    "Health Care",
    "Industrials",
    "Information Technology",
    "Materials",
    "Real Estate",
    "Utilities",
]

EXCHANGES = {
    "S&P 500":  "sp500",
    "NYSE":     "nyse",
    "NASDAQ":   "nasdaq",
}

# Keywords used to detect and exclude non-company securities (ETFs, funds, trusts, etc.)
# These are matched against the security's longName / shortName from yfinance.
ETF_KEYWORDS = [
    "etf", "fund", "index", "trust", "ishares", "invesco", "vanguard",
    "spdr", "proshares", "direxion", "wisdomtree", "vaneck", "schwab",
    "fidelity select", "global x", "ark ", "pimco", "blackrock",
    "portfolio", "income", "bond", "treasury", "commodity", "reit index",
    "preferred", "notes", "debenture", "warrant", "certificate",
]

# ───────────────────────────────────────────────────────────────
# SIDEBAR — FILTERS
# ───────────────────────────────────────────────────────────────
with st.sidebar:
    st.header("⚙️ Filters")

    exchange = st.selectbox(
        "Exchange / Universe",
        options=list(EXCHANGES.keys()),
        index=0,
        help="S&P 500 = ~500 stocks (fast). NYSE/NASDAQ = 2,000–3,500 stocks (slow, 30–60+ min)."
    )

    sector = st.selectbox(
        "Market Sector",
        options=ALL_SECTORS,
        index=0,
        help="Filter to a specific GICS sector. Applied after scan using yfinance data."
    )

    max_price = st.slider(
        "Max Share Price ($)",
        min_value=10, max_value=1000, value=250, step=10,
        help="Exclude stocks above this price."
    )

    min_score = st.slider(
        "Min Hybrid Score",
        min_value=0.0, max_value=20.0, value=2.0, step=0.5,
        help="Only show stocks with a score above this threshold."
    )

    top_n = st.slider(
        "Max Results",
        min_value=5, max_value=100, value=30, step=5,
        help="Maximum number of stocks to display."
    )

    use_ma50_filter = st.toggle(
        "Only stocks BELOW 50-day MA",
        value=True,
        help="When on, only shows stocks in a pullback (price < MA50)."
    )

    st.divider()
    st.header("📊 Metrics  (toggle to include/exclude)")
    metric_enabled = {}
    for key, cfg in METRICS.items():
        metric_enabled[key] = st.toggle(
            cfg["label"],
            value=True,
            help=cfg["desc"],
        )

    st.divider()
    st.header("🔧 Performance")
    max_workers = st.slider(
        "Parallel Workers",
        min_value=1, max_value=20, value=10, step=1,
        help="Higher = faster scan but more network load."
    )

    mfi_period = st.slider(
        "MFI Period (days)",
        min_value=7, max_value=30, value=14, step=1,
        help="Money Flow Index lookback window."
    )

    st.divider()
    run_btn = st.button("🚀 Run Screener", use_container_width=True, type="primary")

# ───────────────────────────────────────────────────────────────
# SIGNAL FUNCTIONS  (identical logic to stock_screener_2026.py)
# ───────────────────────────────────────────────────────────────

def get_buying_pressure(stock, mfi_period):
    try:
        hist = stock.history(period="6mo")
        if hist.empty or len(hist) < mfi_period + 5:
            return 0
        close, high, low, vol = hist["Close"], hist["High"], hist["Low"], hist["Volume"]

        direction = np.sign(close.diff().fillna(0))
        obv = (direction * vol).cumsum()
        obv_slope = np.polyfit(range(20), obv.iloc[-20:].values, 1)[0]
        obv_score = 1.0 if obv_slope > 0 else 0.0

        typical_price = (high + low + close) / 3
        raw_mf = typical_price * vol
        tp_diff = typical_price.diff()
        pos_mf = raw_mf.where(tp_diff > 0, 0).rolling(mfi_period).sum()
        neg_mf = raw_mf.where(tp_diff < 0, 0).rolling(mfi_period).sum()
        mfr = pos_mf / neg_mf.replace(0, np.nan)
        mfi = 100 - (100 / (1 + mfr))
        mfi_latest = mfi.iloc[-1]
        mfi_score = max(0.0, (mfi_latest - 50) / 50) if pd.notnull(mfi_latest) else 0.0

        recent = hist.iloc[-20:].copy()
        recent["up_day"] = recent["Close"] > recent["Close"].shift(1)
        up_vol = recent.loc[recent["up_day"], "Volume"].sum()
        total_vol = recent["Volume"].sum()
        pcv_ratio = up_vol / total_vol if total_vol > 0 else 0.5
        pcv_score = max(0.0, (pcv_ratio - 0.5) / 0.5)

        return round((obv_score + mfi_score + pcv_score) / 3, 4)
    except:
        return 0


def calculate_piotroski(stock):
    try:
        fin, bal, cf = stock.financials, stock.balance_sheet, stock.cashflow
        score = 0
        roa = fin.loc["Net Income"] / bal.loc["Total Assets"]
        if roa.iloc[0] > 0: score += 1
        if cf.loc["Operating Cash Flow"].iloc[0] > 0: score += 1
        if roa.iloc[0] > roa.iloc[1]: score += 1
        if cf.loc["Operating Cash Flow"].iloc[0] > fin.loc["Net Income"].iloc[0]: score += 1
        if bal.loc["Long Term Debt"].iloc[0] < bal.loc["Long Term Debt"].iloc[1]: score += 1
        if (bal.loc["Current Assets"].iloc[0] / bal.loc["Current Liabilities"].iloc[0]) > \
           (bal.loc["Current Assets"].iloc[1] / bal.loc["Current Liabilities"].iloc[1]): score += 1
        return score
    except:
        return None


def get_owner_earnings(stock, info):
    try:
        cf, fin = stock.cashflow, stock.financials
        oe = fin.loc["Net Income"].iloc[0] + cf.loc["Depreciation"].iloc[0] - abs(cf.loc["Capital Expenditure"].iloc[0])
        mc = info.get("marketCap")
        return oe, (oe / mc if mc else None)
    except:
        return None, None


def _get_fin_value(fin, *labels):
    for label in labels:
        if label in fin.index:
            return fin.loc[label]
    return None

def _get_bal_value(bal, *labels):
    for label in labels:
        if label in bal.index:
            return bal.loc[label]
    return None

def calculate_roic(stock):
    try:
        fin, bal = stock.financials, stock.balance_sheet
        ebit_s = _get_fin_value(fin, "EBIT", "Ebit", "Operating Income", "OperatingIncome", "EBITDA", "Ebitda")
        if ebit_s is None: return None
        ebit = ebit_s.iloc[0]
        assets_s = _get_bal_value(bal, "Total Assets", "TotalAssets")
        if assets_s is None: return None
        liab_s = _get_bal_value(bal, "Total Current Liabilities", "TotalCurrentLiabilities", "Current Liabilities", "CurrentLiabilities")
        if liab_s is None: return None
        cash_s = _get_bal_value(bal, "Cash And Cash Equivalents", "Cash Cash Equivalents And Short Term Investments", "CashAndCashEquivalents", "Cash")
        cash = cash_s.iloc[0] if cash_s is not None else 0
        ic = assets_s.iloc[0] - liab_s.iloc[0] - cash
        if pd.isna(ic) or ic == 0: return None
        roic = ebit * 0.79 / ic
        return None if pd.isna(roic) else roic
    except:
        return None


def calculate_roic_trend(stock):
    try:
        fin, bal = stock.financials, stock.balance_sheet
        ebit_s = _get_fin_value(fin, "EBIT", "Ebit", "Operating Income", "OperatingIncome", "EBITDA", "Ebitda")
        if ebit_s is None or len(ebit_s) < 2: return None
        assets_s = _get_bal_value(bal, "Total Assets", "TotalAssets")
        liab_s   = _get_bal_value(bal, "Total Current Liabilities", "TotalCurrentLiabilities", "Current Liabilities", "CurrentLiabilities")
        if assets_s is None or liab_s is None: return None
        if len(assets_s) < 2 or len(liab_s) < 2: return None
        def roic_at(i):
            ic = assets_s.iloc[i] - liab_s.iloc[i]
            if pd.isna(ic) or ic == 0: return None
            v = ebit_s.iloc[i] / ic
            return None if pd.isna(v) else v
        r0, r1 = roic_at(0), roic_at(1)
        return (r0 - r1) if (r0 is not None and r1 is not None) else None
    except:
        return None


def process_ticker(args):
    t, mfi_period = args
    try:
        stock = yf.Ticker(t)
        info  = stock.info

        # Exclude ETFs, index funds, trusts, and other non-company securities
        if is_etf_or_fund(info):
            return None

        price = info.get("currentPrice")
        owner_earnings, oe_yield = get_owner_earnings(stock, info)
        buying_pressure = get_buying_pressure(stock, mfi_period)
        try:
            hist = stock.history(period="3mo")
            ma50 = round(hist["Close"].rolling(50).mean().iloc[-1], 2) if len(hist) >= 50 else None
        except:
            ma50 = None
        return {
            "Ticker":         t,
            "Sector":         info.get("sector", "Unknown"),
            "Price":          price,
            "MA50":           ma50,
            "MarketCap":      info.get("marketCap"),
            "P/E":            info.get("trailingPE"),
            "OwnerEarnings":  owner_earnings,
            "OE_Yield":       oe_yield,
            "ROIC":           calculate_roic(stock),
            "ROIC_Trend":     calculate_roic_trend(stock),
            "RevenueGrowth":  info.get("revenueGrowth"),
            "EarningsGrowth": info.get("earningsGrowth"),
            "Piotroski":      calculate_piotroski(stock),
            "BuyingPressure": buying_pressure,
        }
    except:
        return None

# ───────────────────────────────────────────────────────────────
# TICKER LOADERS
# ───────────────────────────────────────────────────────────────

@st.cache_data(ttl=86400)   # cache for 24 hours so repeated runs are instant
def load_tickers(exchange_key: str) -> list:
    """Return a list of tickers for the chosen exchange."""
    if exchange_key == "sp500":
        url  = "https://en.wikipedia.org/wiki/List_of_S%26P_500_companies"
        resp = requests.get(url, headers={"User-Agent": "Mozilla/5.0"})
        resp.raise_for_status()
        df = pd.read_html(resp.text)[0]
        return df["Symbol"].str.replace(".", "-", regex=False).tolist()

    elif exchange_key == "nyse":
        # NASDAQ's free FTP-style screener API works for NYSE too
        url  = "https://api.nasdaq.com/api/screener/stocks?tableonly=true&limit=5000&exchange=nyse"
        resp = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=30)
        resp.raise_for_status()
        data = resp.json().get("data", {}).get("table", {}).get("rows", [])
        return [r["symbol"].strip() for r in data if r.get("symbol")]

    elif exchange_key == "nasdaq":
        url  = "https://api.nasdaq.com/api/screener/stocks?tableonly=true&limit=5000&exchange=nasdaq"
        resp = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=30)
        resp.raise_for_status()
        data = resp.json().get("data", {}).get("table", {}).get("rows", [])
        return [r["symbol"].strip() for r in data if r.get("symbol")]

    return []


def is_etf_or_fund(info: dict) -> bool:
    """
    Return True if the security is an ETF, index fund, trust, or other
    non-single-company product that should be excluded from the screener.
    Checks yfinance quoteType first (most reliable), then falls back to
    name keyword matching.
    """
    quote_type = (info.get("quoteType") or "").lower()
    if quote_type in ("etf", "mutualfund", "index", "future", "option", "currency", "cryptocurrency"):
        return True

    # Fallback — check name for known fund/ETF keywords
    name = (info.get("longName") or info.get("shortName") or "").lower()
    return any(kw in name for kw in ETF_KEYWORDS)

def build_excel(screened: pd.DataFrame) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        screened.to_excel(writer, index=False, sheet_name="Stock_Picks")
        sheet = writer.book["Stock_Picks"]
        col_map = {cell.value: get_column_letter(cell.column) for cell in sheet[1]}
        for col_name in ["RevenueGrowth", "EarningsGrowth"]:
            if col_name in col_map:
                for cell in sheet[col_map[col_name]][1:]:
                    if cell.value is not None:
                        cell.number_format = "0.00%"
        if "OwnerEarnings" in col_map:
            for cell in sheet[col_map["OwnerEarnings"]][1:]:
                if cell.value is not None:
                    cell.number_format = "$#,##0"
        for idx, col in enumerate(screened.columns, 1):
            col_letter = get_column_letter(idx)
            max_len = max([len(str(c.value)) for c in sheet[col_letter] if c.value] + [len(col)])
            sheet.column_dimensions[col_letter].width = max_len + 2
    return output.getvalue()

# ───────────────────────────────────────────────────────────────
# SCORE COLOR MAP
# ───────────────────────────────────────────────────────────────

def color_score(val):
    try:
        v = float(val)
        if v >= 8:   return "background-color: #1a7a1a; color: white"
        elif v >= 5: return "background-color: #4caf50; color: white"
        elif v >= 3: return "background-color: #ff9800; color: white"
        else:        return "background-color: #f44336; color: white"
    except:
        return ""

# ───────────────────────────────────────────────────────────────
# MAIN RUN LOGIC
# ───────────────────────────────────────────────────────────────

if run_btn:
    # Validate at least one metric is on
    active_metrics = [k for k, v in metric_enabled.items() if v]
    if not active_metrics:
        st.error("Please enable at least one metric before running.")
        st.stop()

    exchange_key  = EXCHANGES[exchange]
    sector_label  = sector if sector != "All Sectors" else "All Sectors"

    # Load tickers for selected exchange (cached for 24hrs)
    with st.spinner(f"Loading {exchange} tickers..."):
        try:
            tickers = load_tickers(exchange_key)
        except Exception as e:
            st.error(f"Failed to load tickers: {e}")
            st.stop()

    if not tickers:
        st.error(f"No tickers returned for {exchange}. Try again later.")
        st.stop()

    if exchange == "S&P 500":
        est_time = "~5–10 min"
    elif exchange == "NYSE":
        est_time = "~30–60 min"
    else:
        est_time = "~40–70 min"

    st.info(
        f"Scanning **{len(tickers)}** tickers on **{exchange}** · "
        f"Sector: **{sector_label}** · ETFs/funds excluded · Est. time: {est_time}"
    )

    # Progress bar + threaded scan
    progress_bar = st.progress(0, text="Starting scan...")
    results = []
    total = len(tickers)

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(process_ticker, (t, mfi_period)): t for t in tickers}
        done = 0
        for future in futures:
            result = future.result()
            if result:
                results.append(result)
            done += 1
            pct = done / total
            progress_bar.progress(pct, text=f"Scanning... {done}/{total} ({int(pct*100)}%)")

    progress_bar.empty()

    if not results:
        st.error("No results returned. Check your internet connection.")
        st.stop()

    # Build DataFrame
    df = pd.DataFrame(results)
    df.replace(["N/A", "None", "-", ""], pd.NA, inplace=True)

    numeric_cols = ["Price", "MA50", "MarketCap", "P/E", "OwnerEarnings", "OE_Yield",
                    "ROIC", "ROIC_Trend", "RevenueGrowth", "EarningsGrowth",
                    "Piotroski", "BuyingPressure"]
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    df["ROIC_Trend"]     = df["ROIC_Trend"].fillna(np.nan)
    df["BuyingPressure"] = df["BuyingPressure"].fillna(0)

    # ── Sector filter — applied here using yfinance sector (authoritative) ──
    if sector != "All Sectors":
        df = df[df["Sector"].str.strip().str.lower() == sector.strip().lower()]
        if df.empty:
            st.error(f"No results found for sector: **{sector}**. The sector name may differ from yfinance labels.")
            st.stop()
    df["P/E"]            = df["P/E"].round(2)
    df["OE_Yield"]       = df["OE_Yield"].round(2)
    df["EarningsYield"]  = df["P/E"].apply(lambda x: round(1/x, 2) if pd.notnull(x) and x > 0 else 0)
    df["ROIC"]           = df["ROIC"].round(2)
    df["ROIC_Trend"]     = df["ROIC_Trend"].round(2)
    df["BuyingPressure"] = df["BuyingPressure"].round(4)
    df["MA50"]           = pd.to_numeric(df["MA50"], errors="coerce").round(2)
    df["Rank_EY"]        = df["EarningsYield"].rank(ascending=False, method="min")
    df["Rank_ROIC"]      = df["ROIC"].fillna(0).rank(ascending=False, method="min")
    df["MagicFormula"]   = df["Rank_EY"] + df["Rank_ROIC"]

    # Dynamic score — only include metrics that are toggled ON
    score = pd.Series(0.0, index=df.index)
    for key, cfg in METRICS.items():
        if metric_enabled.get(key, False):
            score += df[key].fillna(0) * cfg["weight"]
    df["Score"] = score.round(2)

    # Apply filters
    under_price = df["Price"].isna() | (df["Price"] <= max_price)
    above_score = df["Score"] >= min_score
    if use_ma50_filter:
        below_ma50 = df["Price"].isna() | df["MA50"].isna() | (df["Price"] <= df["MA50"])
    else:
        below_ma50 = pd.Series(True, index=df.index)

    screened = (
        df[under_price & below_ma50 & above_score]
        .sort_values("Score", ascending=False)
        .head(top_n)
        .reset_index(drop=True)
    )

    # ── Active metrics summary banner ────────────────────────────
    active_labels = [METRICS[k]["label"] for k in active_metrics]
    st.success(f"**Active metrics:** {' · '.join(active_labels)}")

    # ── Results summary cards ────────────────────────────────────
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Tickers Scanned", len(df))
    col2.metric("Passed Filters",  len(screened))
    col3.metric("Avg Score",       f"{screened['Score'].mean():.2f}" if not screened.empty else "—")
    col4.metric("Top Score",       f"{screened['Score'].max():.2f}"  if not screened.empty else "—")

    st.divider()

    if screened.empty:
        st.warning("No stocks passed the current filters. Try loosening the Score threshold, disabling the MA50 filter, or selecting a different sector.")
    else:
        # Format display columns
        display = screened.copy()
        display["MarketCap"]      = display["MarketCap"].apply(lambda x: f"${x:,.0f}" if pd.notnull(x) else "")
        display["OwnerEarnings"]  = display["OwnerEarnings"].apply(lambda x: f"${x:,.0f}" if pd.notnull(x) else "")
        display["RevenueGrowth"]  = display["RevenueGrowth"].apply(lambda x: f"{x:.2%}" if pd.notnull(x) else "")
        display["EarningsGrowth"] = display["EarningsGrowth"].apply(lambda x: f"{x:.2%}" if pd.notnull(x) else "")

        # Round all remaining numeric columns to 2 decimal places
        for col in display.columns:
            if col not in ["Ticker", "Sector", "MarketCap", "OwnerEarnings", "RevenueGrowth", "EarningsGrowth"]:
                if pd.api.types.is_numeric_dtype(display[col]):
                    display[col] = display[col].apply(lambda x: round(x, 2) if pd.notnull(x) else x)

        # Hide metrics that are toggled off from the display table
        hidden_cols = [k for k, v in metric_enabled.items() if not v and k in display.columns]
        display = display.drop(columns=hidden_cols, errors="ignore")

        styled = display.style.applymap(color_score, subset=["Score"])

        st.subheader(f"Top {len(screened)} Stocks — {sector_label}")
        st.dataframe(styled, use_container_width=True, height=600)

        # Download button
        st.divider()
        excel_bytes = build_excel(screened)
        st.download_button(
            label="⬇️ Download Excel",
            data=excel_bytes,
            file_name=f"stock_screener_{sector.replace(' ', '_')}_{datetime.today().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

else:
    st.info("👈 Configure your filters in the sidebar, then click **Run Screener** to start.")
    st.markdown("""
    ### How to use
    1. **Pick an exchange** — S&P 500 (~500 stocks, fast), NYSE or NASDAQ (2,000–3,500 stocks, 30–70 min)
    2. **Pick a sector** — scan one GICS sector or all sectors
    3. **Toggle metrics on/off** — disabled metrics are excluded from scoring AND hidden from results
    4. **Set your filters** — price cap, min score, MA50 toggle
    5. Hit **🚀 Run Screener**

    > ⚠️ **ETFs, index funds, trusts, and other non-company securities are automatically excluded** from all results.

    ### Metric weights (when enabled)
    | Metric | Weight | Ideal |
    |---|---|---|
    | OE Yield (Owner Earnings / Market Cap) | ×3 | > 5% |
    | ROIC (Return on Invested Capital) | ×2 | > 15% |
    | ROIC Trend (YoY change) | ×2 | Positive |
    | Revenue Growth | ×1 | > 5% |
    | Earnings Growth | ×1 | > 5% |
    | Piotroski Score | ×1 | 7–9 |
    | Buying Pressure (OBV + MFI + PCV) | ×2 | > 0.5 |

    ### Active filters
    - **Exchange** — defines the universe of tickers to scan
    - **Sector** — narrows results to one GICS sector (applied post-scan)
    - **Price cap** — stocks above your max price are excluded
    - **MA50 toggle** — when on, only stocks trading *below* their 50-day MA are shown
    - **Min score** — removes low-conviction picks from the table
    """)
